from django.shortcuts import render, HttpResponse, get_object_or_404
import datetime
import plotly.express as px
from plotly.offline import plot
import pandas as pd
from io import BytesIO

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from extraccion.models import Extraccion
from .forms import DateFilterForm
# Create your views here.

def index(request):
    return HttpResponse("Dashboards")

def tabla_registros(request):
    """
        Muestra los registros de una extraccion.
    """

    form = DateFilterForm(request.GET)

    id_extraccion = request.session.get("id_extraccion")

    if not id_extraccion:
        extraccion_reciente = Extraccion.objects.order_by('-fecha_creacion').first()
        id_extraccion = extraccion_reciente.id


    extraccion = get_object_or_404(Extraccion, pk=id_extraccion)

    registros = extraccion.registro_set.all()

    if form.is_valid():
        fecha_inicio = form.cleaned_data.get('start_date')
        fecha_fin = form.cleaned_data.get('end_date')

        if fecha_inicio:
            # Filtrar objetos donde la fecha sea mayor o igual que la fecha de inicio
            registros = registros.filter(fecha_vencimiento__gte=fecha_inicio)

        if fecha_fin:
            # Filtrar objetos donde la fecha sea menor o igual que la fecha de fin
            queryset = queryset.filter(fecha_vencimiento__lte=fecha_fin)

    context = {
        'form': form,
        'datos': registros,
    }

    return render(request, "dashboard/tabla.html", context)

def mostrar_dashboard(request):
    """
        Muestra graficas sobre las cuentas por cobrar.
    """

    form = DateFilterForm(request.GET)

    id_extraccion = request.session.get("id_extraccion")

    if not id_extraccion:
        extraccion_reciente = Extraccion.objects.order_by('-fecha_creacion').first()
        id_extraccion = extraccion_reciente.id


    extraccion = get_object_or_404(Extraccion, pk=id_extraccion)

    registros = extraccion.registro_set.all()

    # convertimos a un dataframe

    registros_list = list(registros.values())

    df = pd.DataFrame(registros_list)
    df['total'] = df['total'].astype(float)
    df['abonado'] = df['abonado'].astype(float)
    df['debe'] = df['debe'].astype(float)

    # Calcular el saldo pendiente
    df['pendiente'] = df['total'] - df['abonado']

    # Convertir fecha_vencimiento a datetime y calcular días hasta vencimiento
    df['fecha_vencimiento'] = pd.to_datetime(df['fecha_vencimiento'])
    df['dias_hasta_vencimiento'] = (df['fecha_vencimiento'] - pd.Timestamp.now()).dt.days

    # --- GRÁFICO 1: Saldo Pendiente por Proveedor (Barras) ---
    df_tienda_pendiente = df.groupby('tienda')['pendiente'].sum().reset_index()
    fig1 = px.bar(df_tienda_pendiente,
                  x='tienda',
                  y='pendiente',
                  title='Saldo Pendiente por Tienda',
                  labels={'tienda': 'Tienda', 'pendiente': 'Monto Pendiente ($)'},
                  color='tienda' # Colorear por proveedor
    )
    plot_div1 = plot(fig1, output_type='div', include_plotlyjs='cdn', auto_open=False)

    # --- GRÁFICO 2: Distribución del Total Original (Pastel) ---
    fig2 = px.pie(df,
                  names='tienda',
                  values='total',
                  title='Distribución del Total Original de Deuda por Tienda',
                  hole=0.3 # Gráfico de dona
    )
    plot_div2 = plot(fig2, output_type='div', include_plotlyjs='cdn', auto_open=False)

    # --- GRÁFICO 3: Cuentas Pendientes por Vencimiento (Dispersión) ---
    # Filtrar solo las cuentas con saldo pendiente
    df_pendientes = df[df['pendiente'] > 0]
    fig3 = px.scatter(df_pendientes,
                      x='dias_hasta_vencimiento',
                      y='pendiente',
                      size='total', # El tamaño de la burbuja es el total original
                      color='tienda', # Colorear por proveedor
                      hover_name='nombre', # Mostrar ID de cuenta al pasar el ratón
                      title='Cuentas por Cobrar: Días hasta Vencimiento vs. Saldo',
                      labels={'dias_hasta_vencimiento': 'Días hasta Vencimiento (positivo = futuro, negativo = vencido)',
                              'pendiente': 'Saldo Pendiente ($)'},
                      text='nombre' # Mostrar ID de cuenta en el gráfico
    )
    # Ajustar el texto para que se vea mejor (opcional)
    fig3.update_traces(textposition='top center')
    fig3.update_layout(uniformtext_minsize=8, uniformtext_mode='hide')

    plot_div3 = plot(fig3, output_type='div', include_plotlyjs='cdn', auto_open=False)


    context = {
        'chart1_html': plot_div1, # Renombré las variables para mayor claridad
        'chart2_html': plot_div2,
        'chart3_html': plot_div3,
        'page_title': 'Dashboard de Cuentas por Cobrar'
    }

    return render(request, 'dashboard/dashboards.html', context)


def exportar_a_excel(request):
    """
        Convierte el queryset a un archivo de excel.
    """

    id_extraccion = request.session.get("id_extraccion")

    if not id_extraccion:
        extraccion_reciente = Extraccion.objects.order_by('-fecha_creacion').first()
        id_extraccion = extraccion_reciente.id


    extraccion = get_object_or_404(Extraccion, pk=id_extraccion)

    registros = extraccion.registro_set.all().values()

    # Convertir el QuerySet a un DataFrame de Pandas
    df = pd.DataFrame(registros)

    # Crear un archivo Excel en memoria
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, sheet_name='Extraccion', index=False) # index=False para no incluir el índice del DataFrame

    # --- 2. Acceder al workbook y worksheet para más personalización ---
    workbook = writer.book
    sheet = writer.sheets['Extraccion'] # Acceder a la hoja que creamos

    # 2.1. Aplicar estilos a encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid") # Fondo verde
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(sheet[1]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        # Ajustar ancho de columna automáticamente
        max_length = 0
        for i, cell_in_col in enumerate(sheet[get_column_letter(col_idx + 1)]):
            if i == 0: continue # Skip header row for length calculation if preferred
            try:
                if len(str(cell_in_col.value)) > max_length:
                    max_length = len(str(cell_in_col.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        if adjusted_width > 0: # Ensure positive width
            sheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width

    sheet.insert_rows(1, amount=3)

    sheet['D1'] = "CUENTAS POR COBRAR"
    sheet['D1'].font = Font(bold=True, size=16)
    sheet['D1'].alignment = Alignment(horizontal='center', vertical='center')

    sheet['A2'] = "Datos actualizados al:"
    sheet['B2'] = f"{extraccion.fecha_creacion.strftime('%Y-%m-%d %H:%M')}"
    sheet['B2'].alignment = Alignment(horizontal='left')
    sheet['C2'] = "Total de Registros:"
    sheet['D2'] = registros.count()
    sheet['D2'].font = Font(bold=True)

    writer.close()

    # Configurar la respuesta HTTP para descargar el archivo
    output.seek(0) # Mover el puntero al inicio del archivo
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Extraccion.xlsx"'
    return response